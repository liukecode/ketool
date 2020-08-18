#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author  :LiuKe
# @Version : V

import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl


# 获取sheet内容
def get_data(filename, sheet_id):
    data = xlrd.open_workbook(filename)
    table = data.sheets()[sheet_id]
    return table

# 获取行


def get_lines(filename, sheet_id=0):
    table = get_data(filename, sheet_id)
    return table.nrows

# 获取列


def get_cols(filename, sheet_id=0):
    table = get_data(filename, sheet_id)
    return table.ncols

# 获取单元格内容


def get_cells(row, col, filename, sheet_id=0):
    return get_data(filename, sheet_id).cell_value(row, col)
# 获取行数据


def get_row_value(row, filename, sheet_id=0):
    tables = get_data(filename, sheet_id)
    row_data = tables.row_values(row)
    return row_data
# 获取列数据


def get_col_value(col, filename, sheet_id=0):
    tables = get_data(filename, sheet_id)
    col_data = tables.col_values(col)
    return col_data

# 更新excel数据


def update_data(filename, row, col, value):
    if filename.endswith('.xls'):
        wt_update(filename, row, col, value)
    elif filename.endswith('.xlsx'):
        pyxl_update(filename, row, col, value)
    else:
        print('excel file error')

# xls格式表数据更新


def wt_update(filename, row, col, value):
    read_data = xlrd.open_workbook(filename)
    write_data = copy(read_data)
    sheet_data = write_data.get_sheet(0)
    sheet_data.write(row, col, value)
    write_data.save(filename)

# xlsx格式表数据更新


def pyxl_update(filename, row, col, value):
    '''openpyxl 处理数据下标从1开始，已经做了+1处理'''
    wb = openpyxl.load_workbook(filename)  # 生成一个已存在的wookbook对象
    wb1 = wb.active  # 激活sheet
    wb1.cell(row + 1, col + 1, value)  # 写入数据
    wb.save(filename)  # 保存

# 写入excel数据
# kv=[[row,col,value]]


def write_xl(filename, kv):
    if filename.endswith('.xls'):
        wt_write(filename, kv)
    elif filename.endswith('.xlsx'):
        pyxl_write(filename, kv)
    else:
        print('excel file error')
# 写入xls格式excel数据


def wt_write(filename, kv):
    workbook = xlwt.Workbook()  # 创建一个新的工作簿
    sheet = workbook.add_sheet('sheet1')
    for i in kv:
        row, col, value = i
        sheet.write(row, col, value)
    workbook.save(filename)

# 写入xlsx格式excel数据


def pyxl_write(filename, kv):
    # 创建workbook对象，写入模式
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(index=0)
    for i in kv:
        row, col, value = i
        ws.cell(row=row + 1, column=col + 1, value=value)
    wb.save(filename)